#Importing required libraries
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import concurrent.futures
import functools
import time

start = time.perf_counter()
#744


#Function to save text file with URL_ID as its file name.
def save_text_file(url_id, title, content):
    with open(f'{url_id}.txt', 'w', encoding='utf-8') as outfile:
        outfile.write(title + '\n' + content)
        outfile.close()

#Function to Scrape the Article Title and Article Text from the required link
def scrape_text(driver_path, option, class_name, url, url_id):
    driver = webdriver.Chrome(driver_path, options=option)
    driver.get(url)
    post_content = driver.find_element(by=By.CLASS_NAME, value=class_name)
    text  = post_content.text
    save_text_file(url_id, driver.title, text)
    driver.quit()

#Function to fetch the link from Excel File
def get_link_from_excel(filename, urls, ids):
    wb = load_workbook(filename)
    ws = wb.active
    for row in range(2, 172):
        for col in range(2,3):
            char = get_column_letter(col)
            urls.append(ws[char + str(row)].value)
            ids.append(row-1)




#Main Function
if __name__ == "__main__":
    #Required Variables
    filename = 'Input.xlsx'
    urls = []
    ids = []
    options = Options()
    options.headless = True
    driver_path = r'C:\\Users\\Mahi\\Downloads\\chromedriver_win32\\chromedriver.exe' #Change driver path as per your convenience
    class_name = "td-post-content"
    get_link_from_excel(filename, urls, ids)

    #Creating a partial function of "scrape_text" using functools
    scrp_txt = functools.partial(scrape_text, driver_path, options, class_name)

    #Applying Multi-threading to save the text parallely
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
        f = executor.map(scrp_txt, urls, ids)
        print(f)
    #Execution Time
    finish = time.perf_counter()
    print(f'Finished in {finish-start}')


    