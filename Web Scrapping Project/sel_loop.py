from numpy import save
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import concurrent.futures
import functools
import time

start = time.perf_counter()

def save_text_file(url_id, title, content):
    outfile = open(f'{url_id}.txt', 'w')
    outfile.write(title + '\n')
    outfile.write(content)
    outfile.close()

def scrape_text(url, url_id, driver_path, option, class_name):
    driver = webdriver.Chrome(driver_path, options=option)
    driver.get(url)
    post_content = driver.find_element_by_class_name(class_name)
    text  = post_content.text
    save_text_file(url_id, driver.title, text)
    driver.quit()
    # print(driver.title)
    # print(text)

def get_link_from_excel(filename, driver_path, option, class_name):
    wb = load_workbook(filename)
    ws = wb.active
    for row in range(2, 7):
        for col in range(2,3):
            char = get_column_letter(col)
            scrape_text(ws[char + str(row)].value, (row-1), driver_path, option, class_name)
            # print(ws[char + str(row)].value)




if __name__ == "__main__":
    filename = 'input.xlsx'
    urls = []
    ids = []
    options = Options()
    options.headless = True
    driver_path = r'C:\\Users\\Mahi\\Downloads\\chromedriver_win32\\chromedriver.exe'
    class_name = "td-post-content"
    get_link_from_excel(filename, driver_path, options, class_name)
    finish = time.perf_counter()
    print(f'Finished in {finish-start}')
    